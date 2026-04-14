import pandas as pd
import numpy as np
from openpyxl import load_workbook

#读取excel，处理合并单元格
#替换excel文件路径
file_path = "安徽省001.xlsx"
wb = load_workbook(file_path)
ws = wb.active

merged_cells = ws.merged_cells.ranges

for mc in merged_cells:
    #获取合并区域的坐标
    min_col, min_row,max_col,max_row =mc.bounds
    #
    value = ws.cell(row=min_row,column=min_col).value
    #
    ws.unmerge_cells(str(mc))
    for row in range(min_row,max_row+1):
        for col in range(min_col,max_col+1):
            ws.cell(row=row,column=col,value=value)

temp_path ="temp_already unmerged file.xlsx"
wb.save(temp_path)

df_raw = pd.read_excel(temp_path,header=None)
header_row =None
for i in range(min(10,len(df_raw))):
    row = df_raw.iloc[i].astype(str).str.strip()
    if "地区" in row.values and "指标" in row.values and "单位" in row.values:
        header_row = i
        break
if header_row is None:
    raise ValueError("未找到")
df = df_raw.iloc[header_row].reset_index(drop=True)
df.columns =
'''
df = pd.read_excel(temp_path,header=None)

indicators = df.iloc[1:,0].values
units = df.iloc[1:,1].values
regions = df.iloc[0,2:].values
data =df.iloc[1:,2:].values

long_data = []
for i , region in enumerate(regions):
    for j , indicator in enumerate(indicators):
        long_data.append({
            "地区": region,
            "指标": indicator,
            "单位": units[j],
            "数值": data[j,i]
            })
df_long = pd.DataFrame(long_data)
print(df_long)
df_long["地区"] = df_long["地区"].str.strip()
df_long["指标"] = df_long["指标"].str.strip()
df_long["单位"] = df_long["单位"].str.strip()
df_long["数值"] = pd.to_numeric(df_long["数值"],errors="coerce")

df_long = df_long.dropna(subset=["数值"])
print(df_long)
'''
















            
